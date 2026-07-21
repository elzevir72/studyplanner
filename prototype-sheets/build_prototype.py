"""Study Planner Sheets 프로토타입 생성 스크립트.
Google Sheets로 가져오기(import)할 .xlsx 파일을 만든다.
재실행하면 study_planner_prototype.xlsx를 덮어쓴다.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL = PatternFill("solid", fgColor="2B6F5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
DONE_FILL = PatternFill("solid", fgColor="DCEAE3")
MISSING_FILL = PatternFill("solid", fgColor="F2E1C8")

N_USERS = 10  # 최대 참가자 수

wb = openpyxl.Workbook()

# ---------------------------------------------------------------- Config
ws = wb.active
ws.title = "Config"
ws["A1"] = "이름"
ws["B1"] = "활성"
for c in ("A1", "B1"):
    ws[c].font = HEADER_FONT
    ws[c].fill = HEADER_FILL
for i in range(N_USERS):
    row = i + 2
    ws.cell(row=row, column=1, value=f"참가자{i+1}")
    ws.cell(row=row, column=2, value=True)
ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 10
ws.freeze_panes = "A2"
dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ws.add_data_validation(dv_bool)
dv_bool.add(f"B2:B{N_USERS+1}")

# ---------------------------------------------------------------- Entries
ws = wb.create_sheet("Entries")
headers = ["타임스탬프", "이름", "학습일자", "학습수단", "학습내용",
           "학습량_숫자", "학습량_단위", "목표달성", "메모"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
widths = [16, 12, 12, 16, 16, 12, 12, 10, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A2"
dv_goal = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ws.add_data_validation(dv_goal)
dv_goal.add("H2:H1000")

# ---------------------------------------------------------------- Dashboard
ws = wb.create_sheet("Dashboard")
ws["A1"] = "Study Planner — 그룹 대시보드"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:D1")

ws["A3"] = "기준일"
ws["B3"] = "=TODAY()"
ws["A4"] = "주간 시작(월)"
ws["B4"] = "=B3-WEEKDAY(B3,3)"
ws["A5"] = "주간 종료(일)"
ws["B5"] = "=B4+6"
for c in ("A3", "A4", "A5"):
    ws[c].font = LABEL_FONT

# weekly table
ws["A7"] = "이번 주 현황"
ws["A7"].font = LABEL_FONT
weekly_headers = ["이름", "이번 주 기록 수", "목표 달성 수", "상태"]
for i, h in enumerate(weekly_headers, start=1):
    cell = ws.cell(row=8, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

first_row = 9
for i in range(N_USERS):
    r = first_row + i
    cfg_row = i + 2
    ws.cell(row=r, column=1, value=f"=Config!A{cfg_row}")
    ws.cell(row=r, column=2,
             value=f'=COUNTIFS(Entries!$B:$B,$A{r},Entries!$C:$C,">="&$B$4,Entries!$C:$C,"<="&$B$5)')
    ws.cell(row=r, column=3,
             value=f'=COUNTIFS(Entries!$B:$B,$A{r},Entries!$C:$C,">="&$B$4,Entries!$C:$C,"<="&$B$5,Entries!$H:$H,TRUE)')
    ws.cell(row=r, column=4,
             value=f'=IF($A{r}="","",IF(B{r}=0,"⚪ 미기록","✅ 기록함"))')
last_row = first_row + N_USERS - 1

status_range = f"D{first_row}:D{last_row}"
ws.conditional_formatting.add(
    status_range,
    FormulaRule(formula=[f'ISNUMBER(SEARCH("미기록",D{first_row}))'], fill=MISSING_FILL),
)
ws.conditional_formatting.add(
    status_range,
    FormulaRule(formula=[f'ISNUMBER(SEARCH("기록함",D{first_row}))'], fill=DONE_FILL),
)

not_done_row = last_row + 2
ws.cell(row=not_done_row, column=1, value="이번 주 미기록자").font = LABEL_FONT
ws.cell(
    row=not_done_row, column=2,
    value=(f'=TEXTJOIN(", ",TRUE,ARRAYFORMULA(IF(($A${first_row}:$A${last_row}<>"")'
           f'*($B${first_row}:$B${last_row}=0),$A${first_row}:$A${last_row},"")))'),
)

# monthly section
m_start = not_done_row + 3
ws.cell(row=m_start, column=1, value="이번 달 현황").font = LABEL_FONT
ws.cell(row=m_start + 1, column=1, value="이번 달 시작")
ws.cell(row=m_start + 1, column=2, value="=DATE(YEAR($B$3),MONTH($B$3),1)")
ws.cell(row=m_start + 2, column=1, value="이번 달 종료")
ws.cell(row=m_start + 2, column=2, value="=EOMONTH($B$3,0)")
for rr in (m_start + 1, m_start + 2):
    ws.cell(row=rr, column=1).font = LABEL_FONT

m_header_row = m_start + 4
for i, h in enumerate(["이름", "이번 달 기록 수", "목표 달성 수"], start=1):
    cell = ws.cell(row=m_header_row, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

m_first_row = m_header_row + 1
m_start_cell = f"$B${m_start+1}"
m_end_cell = f"$B${m_start+2}"
for i in range(N_USERS):
    r = m_first_row + i
    cfg_row = i + 2
    ws.cell(row=r, column=1, value=f"=Config!A{cfg_row}")
    ws.cell(row=r, column=2,
             value=f'=COUNTIFS(Entries!$B:$B,$A{r},Entries!$C:$C,">="&{m_start_cell},Entries!$C:$C,"<="&{m_end_cell})')
    ws.cell(row=r, column=3,
             value=f'=COUNTIFS(Entries!$B:$B,$A{r},Entries!$C:$C,">="&{m_start_cell},Entries!$C:$C,"<="&{m_end_cell},Entries!$H:$H,TRUE)')
m_last_row = m_first_row + N_USERS - 1

# shared notes feed
feed_row = m_last_row + 3
ws.cell(row=feed_row, column=1, value="최근 공유 메모").font = LABEL_FONT
ws.cell(
    row=feed_row + 1, column=1,
    value='=QUERY(Entries!A:I,"select C,B,I where I != \'\' order by C desc limit 15",1)',
)

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 14

wb.save("study_planner_prototype.xlsx")
print("saved study_planner_prototype.xlsx")
